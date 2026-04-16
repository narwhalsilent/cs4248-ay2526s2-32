import pandas as pd
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment

# Script is at: outputs/evaluation/human_evaluation/
# Go up 3 levels to reach the project root
PROJECT_ROOT = Path(__file__).resolve().parents[3]

CSV_FILES = [
    "outputs/evaluation/bart_full_vanilla/test_predictions.csv",
    "outputs/evaluation/t5_truncated_vanilla/test_predictions.csv",
    "outputs/evaluation/t5_truncated_vanilla_dpo/test_predictions.csv",
]

OUTPUT_DIR = Path(__file__).resolve().parent
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

SAMPLE_SIZE = 100
RANDOM_SEED = 42

for csv_path_str in CSV_FILES:
    csv_path = PROJECT_ROOT / csv_path_str

    if not csv_path.exists():
        print(f"[SKIP] File not found: {csv_path}")
        continue

    df = pd.read_csv(csv_path)

    n = min(SAMPLE_SIZE, len(df))
    sampled = df.sample(n=n, random_state=RANDOM_SEED).reset_index(drop=True)

    model_name = csv_path.parent.name
    output_path = OUTPUT_DIR / f"{model_name}_sample.xlsx"

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        sampled.to_excel(writer, index=False, sheet_name="Sample")

        ws = writer.sheets["Sample"]

        header_fill = PatternFill("solid", start_color="BDD7EE", end_color="BDD7EE")
        header_font = Font(name="Arial", bold=True)

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for col in ws.columns:
            max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

        for row in ws.iter_rows(min_row=2):
            ws.row_dimensions[row[0].row].height = 15

    print(f"[OK] {model_name} -> {output_path}")

print("\nDone. Excel files saved to:", OUTPUT_DIR)